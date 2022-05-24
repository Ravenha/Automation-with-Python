import openpyxl

# load the file and set it to a variable
inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}

# Loop to iterate through the rows. Add +1 because max excludes the last row so add 1 additional to include the last row
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4)

    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier[supplier_name]
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        products_per_supplier[supplier_name] = 1
